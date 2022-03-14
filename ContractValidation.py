from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula import Tokenizer
import pandas as pd
import time
import ValidationFunctions as vf


'''

read_file_Mac = pd.read_csv (r'C:\\Users\\Gert\\Documents\\Development\\excellekes\\final\\MAC.csv', on_bad_lines='skip')
read_file_Mac.to_excel (r'C:\\Users\\Gert\\Documents\\Development\\excellekes\\final\\MAC.xlsx', index = None, header=True)

read_file_Att = pd.read_csv (r'C:\\Users\\Gert\\Documents\\Development\\excellekes\\final\\ATT.csv', on_bad_lines='skip')
read_file_Att.to_excel (r'C:\\Users\\Gert\\Documents\\Development\\excellekes\\final\\ATT.xlsx', index = None, header=True)
'''



path1 = 'FileBucket\\MAC.xlsx'
path2 = 'FileBucket\\ATT.xlsx'
pathdest = 'FileBucket\\RentalContract1.xlsx'
#pathtest = 'FileBucket\\test.xlsx'
#pathtest2 = 'FileBucket\\test2.xlsx'
#pathtestdest = 'FileBucket\\dest.xlsx'

test1 = 'C:\\Users\\Gert\\Documents\\Development\\excellekes\\1.xlsx'
test2 = 'C:\\Users\\Gert\\Documents\\Development\\excellekes\\2.xlsx'
testdest = 'C:\\Users\\Gert\\Documents\\Development\\excellekes\\dest.xlsx'


machAttsheetTitle = 'Machines + Attachments'

	
# 	validation starters
Purpose = "'!B:L,10,FALSE)"
State = "'!B:L,11)"
ObjectType = "'!B:AH,33)"
InFleetStartDate = "'!B:L,8)"
InFleetEndDate = "'!B:L,9)"

valcol_a = 'AddressLine2'
valcol_b = 'ZipCode'
valcol_c = 'City'
valcol_d = 'CountryISO2'
#valcol_e = 'Latitude'
#valcol_f = 'Longitude'
valcol_g = 'EstimatedEndDate'
valcol_h = 'DeliveryEarliest'
valcol_i = 'DeliveryLatest'
valcol_j = 'Purpose'
valcol_k = 'InFleetStartDate'
valcol_l = 'InFleetEndDate'




"""


ValidationRegime = int(input('What kind of validation do you want? 1 = Full, 2 = Combining and validation, 3 = File Validation only\n'))


startGlobalTimer = time.perf_counter()

if ValidationRegime == 1:

	#Pre

	startPreppingTimer = time.perf_counter()
	print('Start Preparing Machine and Attachment data', flush=True)


	vf.addToLastColumn(path1, 'MACHINE')
	vf.addToLastColumn(path2, 'ATTACHMENT')

	finishPreppingTimer = time.perf_counter()
	print(f'Loading finished in {round(finishPreppingTimer-startPreppingTimer, 3)} seconds', flush=True)


	#Combining

	startCombiningTimer = time.perf_counter()
	print('Starting to combine', flush=True)
	vf.combiningWorkbooks(machAttsheetTitle, pathdest, path1, path2)
	finishCombiningTimer = time.perf_counter()
	print(f'Loading finished in {round(finishCombiningTimer-startCombiningTimer, 3)} seconds', flush=True)


	#Loading

	startLoadingTimer = time.perf_counter()
	print('Starting to load', flush=True)
	wb = load_workbook(filename=pathdest)
	wbs0 = wb.worksheets[0]

	finishLoadingTimer = time.perf_counter()
	print(f'Loading finished in {round(finishLoadingTimer-startLoadingTimer, 3)} seconds', flush=True)



	#Comparing


	print('Comparing data', flush=True)
	startCompareTimer = time.perf_counter()

	vf.vlookupToLastCollumn(wbs0, 'Purpose', machAttsheetTitle, Purpose)
	vf.vlookupToLastCollumn(wbs0, 'State', machAttsheetTitle, State)
	vf.vlookupToLastCollumn(wbs0, 'ObjectType', machAttsheetTitle, ObjectType)
	vf.vlookupToLastCollumn(wbs0, 'InFleetStartDate', machAttsheetTitle, InFleetStartDate)
	vf.vlookupToLastCollumn(wbs0, 'InFleetEndDate', machAttsheetTitle, InFleetEndDate)

	finishCompareTimer = time.perf_counter()
	print(f'Loading finished in {round(finishCompareTimer-startCompareTimer, 3)} seconds', flush=True)


	#Validating

	print('Validating', flush=True)
	startValidatingTimer = time.perf_counter()

	vf.searchForBlanks(wb, wbs0, valcol_a)
	vf.searchForBlanks(wb, wbs0, valcol_b)
	vf.searchForBlanks(wb, wbs0, valcol_c)
	vf.searchForBlanks(wb, wbs0, valcol_d)
	#vf.searchForBlanks(wb, wbs0, valcol_e)
	#vf.searchForBlanks(wb, wbs0, valcol_f)
	vf.searchForBlanks(wb, wbs0, valcol_g)
	vf.searchForBlanks(wb, wbs0, valcol_h)
	vf.searchForBlanks(wb, wbs0, valcol_i)
	vf.searchForBlanks(wb, wbs0, valcol_j)
	vf.searchForBlanks(wb, wbs0, valcol_k)
	vf.searchForValues(wb, wbs0, valcol_l)

	finishValidatingTimer = time.perf_counter()
	print(f'Validating finished in {round(finishValidatingTimer-startValidatingTimer, 3)} seconds', flush=True)

	#Saving

	print('Saving', flush=True)

	wbs0.auto_filter.ref = wbs0.dimensions
	wbs0.freeze_panes = 'A2'

	wb.save(pathdest)
	finishLocalTimer = time.perf_counter()
	finishGlobalTimer = time.perf_counter()
	print(f'In total it took {round(finishGlobalTimer-startGlobalTimer, 2)} seconds')

elif ValidationRegime == 2:


	#Combining


	startCombiningTimer = time.perf_counter()
	print('Starting to combine', flush=True)
	vf.combiningWorkbooks(machAttsheetTitle, pathdest, path1, path2)
	finishCombiningTimer = time.perf_counter()
	print(f'Loading finished in {round(finishCombiningTimer-startCombiningTimer, 3)} seconds', flush=True)
	
	#Loading

	startLoadingTimer = time.perf_counter()
	print('Starting to load', flush=True)
	wb = load_workbook(filename=pathdest)
	wbs0 = wb.worksheets[0]

	wbs0.auto_filter.ref = wbs0.dimensions
	wbs0.freeze_panes = 'A2'

	finishLoadingTimer = time.perf_counter()
	print(f'Loading finished in {round(finishLoadingTimer-startLoadingTimer, 3)} seconds', flush=True)



	#Comparing


	print('Comparing data', flush=True)
	startCompareTimer = time.perf_counter()

	vf.vlookupToLastCollumn(wbs0, 'Purpose', machAttsheetTitle, Purpose)
	vf.vlookupToLastCollumn(wbs0, 'State', machAttsheetTitle, State)
	vf.vlookupToLastCollumn(wbs0, 'ObjectType', machAttsheetTitle, ObjectType)
	vf.vlookupToLastCollumn(wbs0, 'InFleetStartDate', machAttsheetTitle, InFleetStartDate)
	vf.vlookupToLastCollumn(wbs0, 'InFleetEndDate', machAttsheetTitle, InFleetEndDate)

	finishCompareTimer = time.perf_counter()
	print(f'Loading finished in {round(finishCompareTimer-startCompareTimer, 3)} seconds', flush=True)


	#Validating

	print('Validating', flush=True)
	startValidatingTimer = time.perf_counter()

	vf.searchForBlanks(wb, wbs0, valcol_a)
	vf.searchForBlanks(wb, wbs0, valcol_b)
	vf.searchForBlanks(wb, wbs0, valcol_c)
	vf.searchForBlanks(wb, wbs0, valcol_d)
	#vf.searchForBlanks(wb, wbs0, valcol_e)
	#vf.searchForBlanks(wb, wbs0, valcol_f)
	vf.searchForBlanks(wb, wbs0, valcol_g)
	vf.searchForBlanks(wb, wbs0, valcol_h)
	vf.searchForBlanks(wb, wbs0, valcol_i)
	vf.searchForBlanks(wb, wbs0, valcol_j)
	vf.searchForBlanks(wb, wbs0, valcol_k)
	vf.searchForValues(wb, wbs0, valcol_l)

	finishValidatingTimer = time.perf_counter()
	print(f'Validating finished in {round(finishValidatingTimer-startValidatingTimer, 3)} seconds', flush=True)

	#Saving

	print('Saving', flush=True)

	wbs0.auto_filter.ref = wbs0.dimensions
	wbs0.freeze_panes = 'A2'

	wb.save(pathdest)
	finishLocalTimer = time.perf_counter()
	finishGlobalTimer = time.perf_counter()
	print(f'In total it took {round(finishGlobalTimer-startGlobalTimer, 2)} seconds')


elif ValidationRegime == 3:

	#Loading

	startLoadingTimer = time.perf_counter()
	print('Starting to load', flush=True)
	wb = load_workbook(filename=pathdest)
	wbs0 = wb.worksheets[0]

	finishLoadingTimer = time.perf_counter()
	print(f'Loading finished in {round(finishLoadingTimer-startLoadingTimer, 3)} seconds', flush=True)



	#Comparing


	print('Comparing data', flush=True)
	startCompareTimer = time.perf_counter()

	vf.vlookupToLastCollumn(wbs0, 'Purpose', machAttsheetTitle, Purpose)
	vf.vlookupToLastCollumn(wbs0, 'State', machAttsheetTitle, State)
	vf.vlookupToLastCollumn(wbs0, 'ObjectType', machAttsheetTitle, ObjectType)
	vf.vlookupToLastCollumn(wbs0, 'InFleetStartDate', machAttsheetTitle, InFleetStartDate)
	vf.vlookupToLastCollumn(wbs0, 'InFleetEndDate', machAttsheetTitle, InFleetEndDate)

	finishCompareTimer = time.perf_counter()
	print(f'Loading finished in {round(finishCompareTimer-startCompareTimer, 3)} seconds', flush=True)


	#Validating

	print('Validating', flush=True)
	startValidatingTimer = time.perf_counter()

	vf.searchForBlanks(wb, wbs0, valcol_a)
	vf.searchForBlanks(wb, wbs0, valcol_b)
	vf.searchForBlanks(wb, wbs0, valcol_c)
	vf.searchForBlanks(wb, wbs0, valcol_d)
	#vf.searchForBlanks(wb, wbs0, valcol_e)
	#vf.searchForBlanks(wb, wbs0, valcol_f)
	vf.searchForBlanks(wb, wbs0, valcol_g)
	vf.searchForBlanks(wb, wbs0, valcol_h)
	vf.searchForBlanks(wb, wbs0, valcol_i)
	vf.searchForBlanks(wb, wbs0, valcol_j)
	vf.searchForBlanks(wb, wbs0, valcol_k)
	vf.searchForValues(wb, wbs0, valcol_l)


	finishValidatingTimer = time.perf_counter()
	print(f'Validating finished in {round(finishValidatingTimer-startValidatingTimer, 3)} seconds', flush=True)

	#Saving

	print('Saving', flush=True)

	wbs0.auto_filter.ref = wbs0.dimensions
	wbs0.freeze_panes = 'A2'

	wb.save(pathdest)
	finishLocalTimer = time.perf_counter()
	finishGlobalTimer = time.perf_counter()
	print(f'In total it took {round(finishGlobalTimer-startGlobalTimer, 2)} seconds')


"""

wb = load_workbook(filename=pathdest)
wbs0 = wb.worksheets[0]

vf.searchForDuplicates(wb, wbs0, 'LoadFile', 'objectId', 'AssignMachineNo')

wb.save(pathdest)
