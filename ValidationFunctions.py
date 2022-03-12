from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter



def addToLastColumn(path, addition):
	wb = load_workbook(filename=path)
	wbs0 = wb.worksheets[0]
	#print(wbs0.max_column)
	char = get_column_letter(wbs0.max_column+1)
	#print(char)
	i = 1
	for row in wbs0:
		wbs0[char + str(i)] = addition
		i = i+1
	wb.save(path) 
	print('Added ' + addition + ' ' + str(wbs0.max_row) + ' times', flush=True)


def vlookupToLastCollumn(sheet, title, sheetname, formula):
	
	#print(sheet.max_column)
	char = get_column_letter(sheet.max_column+1)
	#print(char)
	i = 1
	for row in sheet:
		#sheet[char + str(i)] = "=VLOOKUP(AB2,'Machines + Attachments'!B:S,10,FALSE)"
		sheet[char + str(i)] = "=VLOOKUP(AB" + str(i) + ",'" + sheetname + formula
		i = i+1
	sheet[char + '1'] = title

	

def sheetCopyPaste(ws, destination):
	for row in ws:
		for cell in row:
			destination[cell.coordinate].value = cell.value



def rowMemory(ws):  #produce the list of items in the particular row
        for row in ws.iter_rows(min_row=2):
            yield [cell.value for cell in row]


def combiningWorkbooks(title, destination, header, second):

	wbh = load_workbook(filename=header)
	wsh_0 = wbh.worksheets[0]

	wb2 = load_workbook(filename=second)
	ws2_0 = wb2.worksheets[0]

	wbdestination = load_workbook(filename=destination)
	wsdestination_1 = wbdestination.create_sheet(title)

	
	sheetCopyPaste(wsh_0, wsdestination_1)
	wbdestination.save(destination)
	wbdestination.close()

	wbdestination = load_workbook(filename=destination)
	wsdestination_1 = wbdestination[str(title)]


	list_to_append = list(rowMemory(ws2_0))
	#print(list_to_append)
	for items in list_to_append:
		#print(items)
		wsdestination_1.append(items)
		wsdestination_1.auto_filter.ref = wsdestination_1.dimensions
		wsdestination_1.freeze_panes = 'A2'
			

	wbdestination.save(destination)
	wbh.close()
	wb2.close()
	print('Everything is combined', flush=True)


def searchForBlanks(wb, ws, header):
	ws2 = wb.create_sheet(header + " irregularities")
	for row in ws.iter_rows(min_row=1, max_row=1):
		ws2.append((cell.value for cell in row))
	check = 0


	#Getting the header coordinates to check
	for col in ws.columns:
		column = get_column_letter(col[0].column)
		#print(column)
		for cell in col:
			if str(cell.value) == str(header):
				char = column
				#print(char)

				#checking if the dedicated columm contains a irregularity and than copying the whole row
				for row in ws:
					#value = ws[char + str(row[0].row)].value
					#coll = ws[char + str(row[0].row)]


					#print(str(coll) + '  ' + str(value))

					if ws[char + str(row[0].row)].value == 0 or ws[char + str(row[0].row)].value == None or ws[char + str(row[0].row)].value == " ":

						#needed to add a dumb calculation to remove the sheet again. Could not make it work with searching for A1 as empty cell. 
						check = check + 1
						#print(ws[char+ str(row[0].row)])
						ws2.append((cell.value for cell in row))


#if there is no need we can delete the sheet again.

	if check == 0:
		print('There where no irregularities in ' + header, flush=True)
		wb.remove_sheet(ws2)

	elif check != 0:
		print('There where irregularities in ' + header, flush=True)
		ws2.auto_filter.ref = ws2.dimensions
		ws2.freeze_panes = 'A2' 
		#i could optimise this to insert the headers here but yeah most of the times there will be issues here







